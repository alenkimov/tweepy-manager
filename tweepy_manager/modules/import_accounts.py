import datetime

from better_proxy import parse_proxy_str

import questionary

from sqlalchemy import select
from sqlalchemy.orm import joinedload
import openpyxl

from common.excell import get_xlsx_filepaths, get_worksheets
from common.sqlalchemy.crud import update_or_create

from ..database.models import TwitterAccount, Proxy, Tag
from ..database import AsyncSessionmaker
from ..paths import INPUT_DIR, OUTPUT_DIR
from ..excell import excell


async def select_and_import_table():
    table_filepaths = get_xlsx_filepaths(INPUT_DIR)

    if len(table_filepaths) == 0:
        template_table_filepath = excell.create_empty_table(INPUT_DIR, "template")
        print(f"Created template XLSX table: {template_table_filepath}")
        return
    elif len(table_filepaths) == 1:
        selected_table_filepath = table_filepaths[0]
    else:
        table_filenames = [filepath.name for filepath in table_filepaths]
        selected_table_filename = await questionary.select("Which table?", choices=table_filenames).ask_async()
        selected_table_filepath = INPUT_DIR / selected_table_filename

    worksheets = get_worksheets(selected_table_filepath)

    # TODO Показывать также количество строк в листах
    selected_worksheet_name = await questionary.select("Which worksheet?", choices=worksheets).ask_async()
    selected_worksheet = worksheets[selected_worksheet_name]
    table_data = excell.read_worksheet(selected_worksheet)

    print(f"Loaded {len(table_data)} rows from {selected_table_filepath.name} ({selected_worksheet_name})")
    async with AsyncSessionmaker() as session:
        for twitter_account_data in table_data:
            if twitter_account_data["country_code"]:  # type: str
                twitter_account_data["country_code"] = twitter_account_data["country_code"].lower()

            twitter_id = None
            if twitter_account_data["twitter"]["id"]:
                twitter_id = int(twitter_account_data["twitter"]["id"])
                del twitter_account_data["twitter"]["id"]

            keys = await questionary.checkbox("What to import:", choices=twitter_account_data).ask_async()
            twitter_account_data = {key: twitter_account_data[key] for key in keys}

            if twitter_id:
                twitter_account, created = await update_or_create(
                    session,
                    TwitterAccount,
                    twitter_account_data["twitter"],
                    filter_by={"id": twitter_id},
                )

            else:
                twitter_account = TwitterAccount(**twitter_account_data["twitter"])
                session.add(twitter_account)
                created = True

            if created: print("(NEW!) ", end="")
            print(repr(twitter_account))

            await session.commit()

            if raw_tags := twitter_account_data["tags"]:  # type: str
                print(f"\tTags: {raw_tags}")
                for tag_name in raw_tags.split(","):  # type: str
                    tag_name = tag_name.strip()
                    await session.merge(Tag(twitter_account_id=twitter_account.database_id, tag=tag_name))

            if twitter_account_data["proxy"]:
                parsed_proxy = parse_proxy_str(twitter_account_data["proxy"])
                parsed_proxy["protocol"] = parsed_proxy["protocol"] or "http"
                # TODO Не фильтровать по протоколу
                twitter_account.proxy, created = await update_or_create(session, Proxy, parsed_proxy, filter_by=parsed_proxy)
                if created:
                    print("\t(NEW!) ", end="")
                else:
                    print("\t")
                print(repr(twitter_account.proxy))

            await session.commit()


async def export_to_xlsx():
    """
    Экспортирует все аккаунты из базы данных в Excell-таблицу.

    Поля:
    - status
    - tags (Теги через запятую)
    - proxy (Прокси в формате URL)
    - country_code
    - username
    - id
    - auth_token
    - email
    - email_password
    - password
    - totp_secret
    - backup_code
    """
    async with AsyncSessionmaker() as session:
        accounts = await session.scalars(select(TwitterAccount).options(
            joinedload(TwitterAccount.proxy),
            joinedload(TwitterAccount.user),
        ))
        accounts = list(accounts)

    accounts = sorted(accounts, key=lambda account: account.status)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    now = datetime.datetime.now()
    worksheet.title = now.strftime("%d.%m.%Y")

    # Заголовки столбцов
    worksheet["A1"] = "Status"
    worksheet["B1"] = "Tags"
    worksheet["C1"] = "Proxy"
    worksheet["D1"] = "Country Code"
    worksheet["E1"] = "Username"
    worksheet["F1"] = "ID"
    worksheet["G1"] = "Auth Token"
    worksheet["H1"] = "Email"
    worksheet["I1"] = "Email Password"
    worksheet["J1"] = "Password"
    worksheet["K1"] = "TOTP Secret"
    worksheet["L1"] = "Backup Code"

    # Заполнение данных
    for row, account in enumerate(accounts, start=2):
        worksheet.cell(row=row, column=1, value=account.status.value)

        tags = await session.scalars(select(Tag).where(Tag.twitter_account_id == account.database_id))
        tags = ", ".join([tag.tag for tag in tags])
        worksheet.cell(row=row, column=2, value=tags)

        if account.proxy:
            proxy = account.proxy.better_proxy()
            worksheet.cell(row=row, column=3, value=str(proxy))
        else:
            worksheet.cell(row=row, column=3, value="")

        worksheet.cell(row=row, column=4, value=account.country_code)
        worksheet.cell(row=row, column=5, value=account.username)
        worksheet.cell(row=row, column=6, value=account.user.id if account.user else "")
        worksheet.cell(row=row, column=7, value=account.auth_token)
        worksheet.cell(row=row, column=8, value=account.email)
        worksheet.cell(row=row, column=9, value=account.email_password)
        worksheet.cell(row=row, column=10, value=account.password)
        worksheet.cell(row=row, column=11, value=account.totp_secret)
        worksheet.cell(row=row, column=12, value=account.backup_code)

    export_dir = OUTPUT_DIR / "export"
    export_dir.mkdir(exist_ok=True)
    workbook.save(export_dir / f"twitter_accounts_{now.strftime('date_%d_%m_%Y.time_%H_%M_%S')}.xlsx")
