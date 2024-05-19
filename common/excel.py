from datetime import datetime
from typing import Any
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


RED = "90EE90"


class Column:

    def __init__(
            self,
            header: str,
            description: str,
            key: str,
            *,
            group_name: str = None,
            required: bool = False,
            max_length: int = None
    ) -> None:
        self.header = header
        self.description = description
        self.key = key
        self.group_name = group_name
        self.required = required
        self.max_length: int = max_length

    @property
    def color(self) -> PatternFill | None:
        if self.required:
            return PatternFill(
                start_color=RED,
                end_color=RED,
                fill_type="solid",
            )

        return None

    @property
    def full_header(self) -> str:
        header = f"[{self.group_name}] " if self.group_name else ""
        header += self.header
        return header

    @property
    def full_description(self) -> str:
        description = "REQUIRED FIELD" if self.required else "OPTIONAL FIELD"
        if self.max_length:
            description += f"\nMAX LENGTH: {self.max_length}"
        description += f"\n\n{self.description}"
        return description


class Excel:
    def __init__(self, columns: list[Column]) -> None:
        self.columns = columns

    def create_empty_table(self, dirpath: Path, name: str) -> Path:
        """Создает пустую таблицу с заголовками и описаниями."""
        wb = Workbook()
        ws = wb.active

        for i, column in enumerate(self.columns, start=1):
            ws.cell(row=1, column=i, value=column.full_header)

        for i, column in enumerate(self.columns, start=1):
            cell = ws.cell(row=2, column=i, value=column.full_description)

            if column.color:
                cell.fill = column.color

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        table_filepath = dirpath / f"{name}.xlsx"
        wb.save(table_filepath)
        return table_filepath

    def read_worksheet(self, worksheet: Worksheet) -> list[dict[str, Any]]:
        """Читает данные из таблицы начиная с третьей строки."""
        # Создаем словарь для сопоставления заголовков столбцов из файла с объектами Column
        column_map: dict[int, Column] = {}
        for i, cell in enumerate(worksheet[1], start=1):  # Предполагаем, что заголовки находятся в первой строке
            for column in self.columns:
                if cell.value == column.full_header:
                    column_map[i] = column
                    break

        data: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            row_data: dict[str, Any] = {}
            for i, value in enumerate(row, start=1):
                column = column_map.get(i)
                if column:
                    if column.group_name:
                        if column.group_name not in row_data:
                            row_data[column.group_name] = {}
                        row_data[column.group_name][column.key] = value
                    else:
                        row_data[column.key] = value
            if row_data:  # Добавляем только если есть данные
                data.append(row_data)

        return data

    def export(self, filepath: Path, data: list[dict[str, Any]]):
        """
        Выгружает рабочую книгу по указанному пути.

        :param filepath: Путь к файлу рабочей книги.
        :param data: Список словарей, где ключи — названия полей, а значения — данные.
        """
        workbook = Workbook()
        worksheet = workbook.active

        now = datetime.now()
        worksheet.title = now.strftime("%d.%m.%Y")

        # Записываем заголовки
        for column_idx, column in enumerate(self.columns, start=1):
            worksheet.cell(row=1, column=column_idx, value=column.full_header)

        # Записываем данные
        for row_idx, row_data in enumerate(data, start=2):
            for column_idx, column in enumerate(self.columns, start=1):
                if column.group_name:
                    group_data = row_data.get(column.group_name, {})
                    value = group_data.get(column.key, "")
                else:
                    value = row_data.get(column.key, "")
                worksheet.cell(row=row_idx, column=column_idx, value=value)

        # Устанавливаем ширину столбцов
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        workbook.save(filepath)


def get_xlsx_filepaths(dirpath: Path) -> list[Path]:
    return list(dirpath.glob("*.xlsx"))


def get_worksheets(filepath: Path) -> dict[str, Worksheet]:
    """
    Загружает рабочую книгу из указанного файла и возвращает словарь её листов,
    где ключи — это названия листов, а значения — объекты листов.

    :param filepath: Путь к файлу рабочей книги.
    :return: Словарь, где ключи — названия листов, а значения — листы.
    """
    workbook = load_workbook(filepath)
    return {sheet.title: sheet for sheet in workbook.worksheets}
