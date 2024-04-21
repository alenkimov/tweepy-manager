from typing import Any, List, Dict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Column:
    DESCRIPTION_CELL_COLOR_IF_REQUIRED = "90EE90"

    def __init__(
            self,
            header: str,
            description: str,
            key: str,
            *,
            group_name: str = None,
            required: bool = False,
            max_length: int = None
    ) -> None:
        self.header = header
        self.description = description
        self.key = key
        self.group_name = group_name
        self.required = required
        self.max_length: int = max_length

    @property
    def color(self):
        if self.required:
            return PatternFill(
                start_color=self.DESCRIPTION_CELL_COLOR_IF_REQUIRED,
                end_color=self.DESCRIPTION_CELL_COLOR_IF_REQUIRED,
                fill_type="solid",
            )

        return None

    @property
    def full_header(self) -> str:
        header = f"[{self.group_name}] " if self.group_name else ""
        header += self.header
        return header

    @property
    def full_description(self) -> str:
        description = "REQUIRED FIELD" if self.required else "OPTIONAL FIELD"
        if self.max_length:
            description += f"\nMAX LENGTH: {self.max_length}"
        description += f"\n\n{self.description}"
        return description


class Excel:
    def __init__(self, columns: list[Column]) -> None:
        self.columns = columns

    def create_empty_table(self, dirpath: Path, name: str) -> Path:
        """Создает пустую таблицу с заголовками и описаниями."""
        wb = Workbook()
        ws = wb.active

        for i, column in enumerate(self.columns, start=1):
            ws.cell(row=1, column=i, value=column.full_header)

        for i, column in enumerate(self.columns, start=1):
            cell = ws.cell(row=2, column=i, value=column.full_description)

            if column.color:
                cell.fill = column.color

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        table_filepath = dirpath / f"{name}.xlsx"
        wb.save(table_filepath)
        return table_filepath

    def read_worksheet(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Читает данные из таблицы начиная с третьей строки."""
        # Создаем словарь для сопоставления заголовков столбцов из файла с объектами Column
        column_map: Dict[int, Column] = {}
        for i, cell in enumerate(worksheet[1], start=1):  # Предполагаем, что заголовки находятся в первой строке
            for column in self.columns:
                if cell.value == column.full_header:
                    column_map[i] = column
                    break

        data: List[Dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            row_data: Dict[str, Any] = {}
            for i, value in enumerate(row, start=1):
                column = column_map.get(i)
                if column:
                    if column.group_name:
                        if column.group_name not in row_data:
                            row_data[column.group_name] = {}
                        row_data[column.group_name][column.key] = value
                    else:
                        row_data[column.key] = value
            if row_data:  # Добавляем только если есть данные
                data.append(row_data)

        return data


def get_xlsx_filepaths(dirpath: Path) -> list[Path]:
    return list(dirpath.glob("*.xlsx"))


def get_worksheets(filepath: Path) -> dict[str, Worksheet]:
    """
    Загружает рабочую книгу из указанного файла и возвращает словарь её листов,
    где ключи — это названия листов, а значения — объекты листов.

    :param filepath: Путь к файлу рабочей книги.
    :return: Словарь, где ключи — названия листов, а значения — листы.
    """
    workbook = load_workbook(filepath)
    return {sheet.title: sheet for sheet in workbook.worksheets}
